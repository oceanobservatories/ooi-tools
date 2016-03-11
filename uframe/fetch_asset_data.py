#!/usr/bin/env python

import os
import click
import datetime
from openpyxl import Workbook

from uframe import SensorInventory, AssetManagement

wanted = ['ADCP', 'BOTPT', 'CAMDS', 'CTD', 'FLOR', 'NUTNR', 'PARAD', 'PREST', 'TMPSF',
          'SPKIR', 'VEL', 'FLCDR', 'FLTNU', 'HPIES', 'HYD', 'OBS', 'ZPLSC', 'DO', 'MASS',
          'D1000', 'PCO', 'PH', 'OPT']

outdir = '/Users/petercable/Google Drive/OOI Driver Development Documents/Calibration/Retrieved'

CALSHEET = 'Asset_Cal_Info!A:H'
external_sheets = {
    'RS01SBPS': 'https://docs.google.com/spreadsheets/d/1XYrEu4eFb_c0ainr4IR6pzHHRFiGEZYFdVp5iD4d8FQ',
    'RS01SLBS': 'https://docs.google.com/spreadsheets/d/1nWNDsOyqrKfeHuUjylKUwecKv6N7UudOCIGiS92lFGY',
    'RS01SBPD': 'https://docs.google.com/spreadsheets/d/1QURqXZ3jmXo6zzebCBK2GXo0LzBR5Cs0GxlG_HjLWM0',
    'RS01SUM2': 'https://docs.google.com/spreadsheets/d/1fPnFsHDj5RcTvdKctmNsdBJAKEE73ASaw_eIIc3jOnc',
    'RS03ASHS': 'https://docs.google.com/spreadsheets/d/1qfI4ix6UdNj24a_8cy6iWrx_sxQpRJJpL2SvPu5e7L0',
    'CE04OSBP': 'https://docs.google.com/spreadsheets/d/1AL4SD3uEuUc71uXXHlOSeuMKvgknfmHdUFcKcprqzxI',
    'CE02SHBP': 'https://docs.google.com/spreadsheets/d/1APRW-aZacbxftlLihb6dBoQx3ziaSdzcy3UJ-YfJR_s',
    'CE04OSPS': 'https://docs.google.com/spreadsheets/d/1i_OJoOG5mApskCZazw71z7niabuzr5OBuldYU0Knb20',
    'CE04OSPD': 'https://docs.google.com/spreadsheets/d/1Twm_jC8VTwYXuA4TF_5UBhunc9FnpI1sbH3ckc3xBT8',
    'RS01SUM1': 'https://docs.google.com/spreadsheets/d/1zI_t0VfE77ETas5glITlAMV-uBW-dvln1RBJfsOFsHs',
    'RS03AXBS': 'https://docs.google.com/spreadsheets/d/1ODaqC0lYOqc6yPodnLrF7JsqNBGBgoxajRCURONG85E',
    'RS03AXPD': 'https://docs.google.com/spreadsheets/d/1caK46PNxE6hWT0XDvnyD0NA3SLkgVhKSlxsXbLZFlew',
    'RS03AXPS': 'https://docs.google.com/spreadsheets/d/1pT_twjCzytNfcsN5A-tqpC-Oh6jBi03-liwOIrHGL3I',
    'RS03CCAL': 'https://docs.google.com/spreadsheets/d/1zLwGealGtLFeZWP2DqQHMLauvIwiePWl-easM53sxuM',
    'RS03ECAL': 'https://docs.google.com/spreadsheets/d/1n-DodehvVoGb5vrWM4BPEcpihLsE51jxCo8D0LuovUQ',
    'RS03INT1': 'https://docs.google.com/spreadsheets/d/11Aye9uGQy0yzb20wDNO7dgFiugcY8PfwaRyBRqEdkTw',
    'RS03INT2': 'https://docs.google.com/spreadsheets/d/1pghVRD4z8J_MYUouQrIkYwyLNkyIc50NFPUSbIzjabw',
}


@click.command()
@click.option('--subsites', default=None, help='One or more subsites to be queried', multiple=True)
@click.option('--nodes', default=None, help='One or more nodes to be queried', multiple=True)
def check_stream(subsites, nodes):
    now = datetime.datetime.utcnow()
    inv = SensorInventory()
    am = AssetManagement()
    deps = []
    cals = []
    if subsites or nodes:
        sensors = inv.get_sensors(subsites=subsites, nodes=nodes)
        for subsite, node, sensor in sensors:
            if node == 'XX00X':
                continue
            for each in wanted:
                if each in sensor:
                    d = am.get_deployments_single(subsite, node, sensor)
                    c = am.get_calibrations_single(subsite, node, sensor)
                    if d:
                        deps.extend(d)
                    else:
                        deps.append(('', '-'.join((subsite, node, sensor))))

                    if c:
                        cals.extend(c)
                    else:
                        cals.append(('', '-'.join((subsite, node, sensor))))
                    break
    else:
        deps = am.get_deployments_all()
        cals = am.get_calibrations_all()

    dep_cols = ['tag', 'refdes', 'deployment', 'start', 'stop']
    cal_cols = dep_cols + ['name', 'value']

    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Deployments'
    sheet.append(dep_cols)
    for row in deps:
        sheet.append(row)
    sheet['H1'] = '''=QUERY(B:C, "select B, count(C) where B != '' group by B")'''

    sheet = wb.create_sheet(title="Calibrations")
    sheet.append(cal_cols)
    for row in cals:
        sheet.append(row)
    sheet['I1'] = '''=QUERY(B:F, "select B, count(F) where B != '' group by B order by B", 1)'''

    for sheetname in external_sheets:
        sheet = wb.create_sheet(title=sheetname)
        sheet['A1'] = '''=IMPORTRANGE("%s", "%s")''' % (external_sheets[sheetname], CALSHEET)
        sheet['K1'] = 'From AM'
        sheet['N1'] = 'From CalSheet'
        sheet['K3'] = '''=QUERY(Calibrations!B:F, "select B,C,count(F)''' +\
                      ''' where B starts with '%s' and F != '' group by B,C order by B,C", 1)''' % sheetname
        sheet['N3'] = '''=QUERY(A:G, "select A,D,count(G) where G != '' group by A,D order by A,D")'''

    fname = os.path.join(outdir, 'DepsCals_%s.xlsx' % now.strftime('%Y%m%d-%H%M'))
    wb.save(fname)

if __name__ == '__main__':
    check_stream()
